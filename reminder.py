from openpyxl import *
import datetime 
import assist
from plyer import notification
def add_remind():
    wb = load_workbook("/Users/apple/Desktop/Book1.xlsx")
    ws=wb["Sheet1"]   
    row_count = ws.max_row
    i = row_count + 1
    print("what is the reminder about")
    assist.speak("what is the reminder about")
    about = assist.get_audio()
    wcell1 = ws.cell(i,1)
    wcell1.value=about
    print("When should i remind you this?")
    assist.speak("When should i remind you this?")
    date = assist.get_audio()
    d = assist.get_date(date)
    wcell2 = ws.cell(i,2)
    wcell2.value=d
    print("plz enter time")
    tt= input()
    print("plz enter time")
    assist.speak("plz enter time")
    wcell3 = ws.cell(i,3)
    wcell3.value=tt
    wb.save("/Users/apple/Desktop/Book1.xlsx")


def read_reminder():
    wb = load_workbook("/Users/apple/Desktop/Book1.xlsx")
    ws=wb["Sheet1"] 
    now = datetime.datetime.now()
    td = now.strftime('%Y-%m-%d')
    tnow = now.strftime('%H:%M:%S')
    row_count = ws.max_row
    print("Reminders for today")
    assist.speak("Reminders for today")
    for i in range(2,row_count+1):
      val = ws.cell(i,2)
      d=val.value
      val1 = ws.cell(i,3)
      d1 =val1.value
      t = datetime.datetime.strftime(d,'%Y-%m-%d')
      t1 = d1.strftime('%H:%M:%S')
      
      
      if(td==t):
        if(tnow<=t1):
         ab =ws.cell(i,1)
         about =ab.value
        
         print(about + " at " + t1)

         
         assist.speak(about + " at " + t1)

         notification.notify(title='Reminder ' ,message=about + " at " + t1,app_icon=None,  timeout=10, )
      

 
      





